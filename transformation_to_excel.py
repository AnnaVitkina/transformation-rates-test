"""
Convert extracted JSON data to multi-tab Excel format for analysis.
Creates one tab per extracted field block (MainCosts, AddedRates, CountryZoning, etc.).

Input:  processing/extracted_data.json  (produced by the extraction pipeline)
Output: output/DHL_Rate_Cards.xlsx      (multi-tab workbook)

HOW THIS FILE FITS INTO THE BIGGER PICTURE
-------------------------------------------
Before this script runs, another script has already read a DHL rate-card PDF and
saved all the pricing data into a single JSON file (extracted_data.json).
This script's job is to take that JSON file and turn it into a nicely formatted
Excel workbook with one tab per data section, so analysts can open it directly.

This file is the "conductor" – it imports the four specialist modules and calls
them in the right order to build the complete workbook:

  transform_main_costs.py   – builds the MainCosts lane matrix
  transform_other_tabs.py   – builds AddedRates, CountryZoning, and other flat tabs
  accessorial_costs.py      – builds the Accessorial Costs tab with fuzzy cost-type matching
  excel_helpers.py          – writes all tabs to the Excel file with formatting

HOW TO RUN:
  python transformation_to_excel.py
  (reads processing/extracted_data.json, writes output/DHL_Rate_Cards.xlsx)

For use from pipeline_main.py, import save_to_excel() directly.
"""

import json
import os
from pathlib import Path

# Import the four specialist modules
from transform_main_costs import build_matrix_main_costs, expand_main_costs_lanes_by_zoning, apply_zone_labels_to_main_costs
from transform_other_tabs import flatten_array_data, pivot_added_rates, build_zone_label_lookup
from accessorial_costs import build_accessorial_costs_rows
from excel_helpers import (
    write_matrix_sheet,
    write_sheet,
    write_accessorial_sheet,
    ACCESSORIAL_COSTS_COLUMNS,
)


# ---------------------------------------------------------------------------
# I/O helper
# ---------------------------------------------------------------------------

def load_extracted_data(filepath):
    """
    Open the extracted JSON file from disk and return its contents as a Python dictionary.
    If the file cannot be opened or is not valid JSON, an error is printed and the
    program stops immediately.
    """
    print(f"[*] Loading extracted data from: {filepath}")
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print(f"[OK] Data loaded successfully")
        return data
    except Exception as e:
        print(f"[ERROR] Failed to load data: {e}")
        raise


# ---------------------------------------------------------------------------
# Metadata tab
# ---------------------------------------------------------------------------

def create_metadata_sheet(workbook, metadata):
    """
    Create the first tab in the Excel file called "Metadata".
    This tab shows basic information about the rate card document:
    who the client is, which carrier it belongs to, when it is valid, etc.
    It is a simple two-column table: column A = field name, column B = value.
    """
    print("[*] Creating Metadata tab...")

    from openpyxl.styles import Font, PatternFill, Alignment

    ws = workbook.create_sheet("Metadata", 0)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    def _str(v):
        s = "" if v is None else (v.replace("\n", " ") if isinstance(v, str) else str(v))
        return s

    data = [
        ["Field", "Value"],
        ["Client", _str(metadata.get("client"))],
        ["Carrier", _str(metadata.get("carrier"))],
        ["Validity Date", _str(metadata.get("validity_date"))],
        ["FileName", _str(metadata.get("FileName"))],
        ["Extraction Date", _str(metadata.get("extraction_date"))],
        ["Extraction Source", _str(metadata.get("extraction_source"))],
    ]

    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60

    print(f"[OK] Metadata tab created")


# ---------------------------------------------------------------------------
# Workbook orchestrator
# ---------------------------------------------------------------------------

def save_to_excel(data, output_path, accessorial_folder=None):
    """
    The main function that builds the complete Excel workbook and saves it to disk.

    This function is the "conductor" – it calls all the other functions in the right
    order and assembles their output into a single multi-tab Excel file.

    TABS CREATED (in order):
      1  Metadata              – basic info: client, carrier, validity date, filename
      2  MainCosts             – the main pricing table (one row per lane, all cost
                                 categories merged, letter zones expanded to real O/D pairs)
      3  AddedRates            – supplemental rate tables (e.g. fuel surcharge by weight/zone)
      4  AdditionalCostsPart1  – first batch of extra charges from the rate card
      5  CountryZoning         – which countries belong to which zone (with ISO codes added)
      6  AdditionalZoning      – additional zoning rules (if present)
      7  GoGreenPlusCost       – GoGreen Plus rows; Origin/Destination lists → DHL codes
      8  ZoningMatrix          – the raw origin/destination zone matrix (for reference)
      9  AdditionalCostsPart2  – second batch of extra charges from the rate card
      10 Accessorial Costs     – combined view of Part1 + Part2 with standardised Cost Types

    PARAMETERS:
      data               – the full JSON dictionary loaded from extracted_data.json
      output_path        – where to save the .xlsx file
      accessorial_folder – folder containing client-specific reference files for Cost Type matching

    RETURNS: the path of the accessorial reference file used (or None if none was found)
    """
    print(f"[*] Creating Excel file: {output_path}")

    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl not installed!")
        print("        To install: pip install openpyxl")
        raise

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove the default empty sheet

        metadata = data.get('metadata', {})

        # -----------------------------------------------------------------------
        # Tab 1: Metadata
        # -----------------------------------------------------------------------
        create_metadata_sheet(wb, metadata)

        # -----------------------------------------------------------------------
        # Tab 2: MainCosts
        # Step A: build_matrix_main_costs() merges all cost categories into one row per lane
        # Step B: expand_main_costs_lanes_by_zoning() replaces letter zones (A, B …)
        #         with real Origin/Destination pairs from the ZoningMatrix
        # -----------------------------------------------------------------------
        main_costs_data = data.get('MainCosts', [])
        zoning_matrix = data.get('ZoningMatrix', [])
        country_zoning = data.get('CountryZoning', [])
        if main_costs_data:
            # Pass zoning_matrix so build_matrix_main_costs can detect matrix zones accurately
            matrix_rows, category_specs = build_matrix_main_costs(main_costs_data, metadata, zoning_matrix)
            if zoning_matrix:
                matrix_rows = expand_main_costs_lanes_by_zoning(matrix_rows, zoning_matrix)
            # Replace raw "Zone 8" style values in Origin/Destination with short labels
            # like "ECONOMY_EXP_ZONE_8" derived from the CountryZoning rate names
            if country_zoning:
                zone_label_lookup = build_zone_label_lookup(country_zoning)
                matrix_rows = apply_zone_labels_to_main_costs(matrix_rows, zone_label_lookup)
            write_matrix_sheet(wb, "MainCosts", matrix_rows, category_specs, metadata)

        # -----------------------------------------------------------------------
        # Tab 3: AddedRates
        # The JSON has interleaved header and data rows; pivot_added_rates() untangles them.
        # -----------------------------------------------------------------------
        added_rates = data.get('AddedRates', [])
        if added_rates:
            added_rates_rows = pivot_added_rates(added_rates, metadata)
            write_sheet(wb, "AddedRates", added_rates_rows, metadata)

        # -----------------------------------------------------------------------
        # Tab 4: AdditionalCostsPart1
        # flatten_array_data() just prepends the three identity columns.
        # -----------------------------------------------------------------------
        additional_costs_1 = data.get('AdditionalCostsPart1', [])
        if additional_costs_1:
            additional_costs_1_rows = flatten_array_data(additional_costs_1, metadata, 'AdditionalCostsPart1')
            write_sheet(wb, "AdditionalCostsPart1", additional_costs_1_rows, metadata)

        # -----------------------------------------------------------------------
        # Tab 5: CountryZoning
        # flatten_array_data() applies two extra enrichment steps for this tab:
        #   - Forward-fill empty RateName cells
        #   - Add a Country Code column (ISO 2-letter codes)
        # -----------------------------------------------------------------------
        country_zoning = data.get('CountryZoning', [])
        if country_zoning:
            country_zoning_rows = flatten_array_data(country_zoning, metadata, 'CountryZoning')
            write_sheet(wb, "CountryZoning", country_zoning_rows, metadata)

        # -----------------------------------------------------------------------
        # Tab 6: AdditionalZoning
        # -----------------------------------------------------------------------
        additional_zoning = data.get('AdditionalZoning', [])
        if additional_zoning:
            additional_zoning_rows = flatten_array_data(additional_zoning, metadata, 'AdditionalZoning')
            write_sheet(wb, "AdditionalZoning", additional_zoning_rows, metadata)

        # -----------------------------------------------------------------------
        # Tab 7: GoGreenPlusCost
        # Origin/Destination: comma-separated "CODE - Name" lists → codes via dhl_country_codes.txt
        # -----------------------------------------------------------------------
        gogreen_plus = data.get('GoGreenPlusCost', [])
        if gogreen_plus:
            gogreen_rows = flatten_array_data(gogreen_plus, metadata, 'GoGreenPlusCost')
            write_sheet(wb, "GoGreenPlusCost", gogreen_rows, metadata)

        # -----------------------------------------------------------------------
        # Tab 8: ZoningMatrix
        # -----------------------------------------------------------------------
        zoning_matrix = data.get('ZoningMatrix', [])
        if zoning_matrix:
            zoning_matrix_rows = flatten_array_data(zoning_matrix, metadata, 'ZoningMatrix')
            write_sheet(wb, "ZoningMatrix", zoning_matrix_rows, metadata)

        # -----------------------------------------------------------------------
        # Tab 9: AdditionalCostsPart2
        # -----------------------------------------------------------------------
        additional_costs_2 = data.get('AdditionalCostsPart2', [])
        if additional_costs_2:
            additional_costs_2_rows = flatten_array_data(additional_costs_2, metadata, 'AdditionalCostsPart2')
            write_sheet(wb, "AdditionalCostsPart2", additional_costs_2_rows, metadata)

        # -----------------------------------------------------------------------
        # Tab 10: Accessorial Costs
        # Combines Part1 and Part2 into one clean table with standardised Cost Types.
        # -----------------------------------------------------------------------
        accessorial_rows, accessorial_file_used = build_accessorial_costs_rows(
            data.get('AdditionalCostsPart1', []),
            data.get('AdditionalCostsPart2', []),
            metadata,
            accessorial_folder=accessorial_folder,
        )
        if accessorial_rows:
            write_accessorial_sheet(wb, "Accessorial Costs", accessorial_rows)

        # Save the finished workbook
        wb.save(output_path)

        # -----------------------------------------------------------------------
        # Post-processing: expand MainCosts (carrier country -> ISO, optional AdditionalZoning).
        # Always run so that Origin/Destination country names (e.g. Switzerland) are
        # converted to 2-letter codes in Origin Country / Destination Country.
        # When AdditionalZoning data is present, also adds rows and columns for
        # starred-country sub-zones (Origin Country, Origin City, etc.).
        # -----------------------------------------------------------------------
        try:
            from expand_additional_zoning import expand_main_costs_with_additional_zoning
            expand_main_costs_with_additional_zoning(output_path)
        except Exception as e:
            print(f"[WARN] MainCosts post-processing failed (non-fatal): {e}")

        file_size = os.path.getsize(output_path)
        file_size_kb = file_size / 1024

        print(f"[OK] Excel file saved successfully")
        print(f"  - Tabs: {len(wb.sheetnames)}")
        print(f"  - File size: {file_size_kb:.2f} KB")

        return str(accessorial_file_used) if accessorial_file_used else None

    except Exception as e:
        print(f"[ERROR] Failed to save Excel: {e}")
        raise


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    """
    The starting point when this script is run directly from the command line.

    This function:
      1. Defines the input and output file paths.
      2. Makes sure the output folder exists (creates it if needed).
      3. Loads the extracted JSON data from disk.
      4. Calls save_to_excel() to build and save the Excel workbook.
      5. Prints a summary of what was created.

    If anything goes wrong at any step, an error message is printed and the
    program stops with a non-zero exit code.
    """
    print("=" * 60)
    print("DHL RATE CARD EXCEL GENERATOR")
    print("=" * 60)
    print()

    input_file = 'processing/extracted_data.json'
    output_dir = 'output'
    output_file = os.path.join(output_dir, 'DHL_Rate_Cards.xlsx')

    try:
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        print(f"[OK] Output directory ready: {output_dir}")
        print()

        print("Step 1: Loading extracted data...")
        data = load_extracted_data(input_file)
        print()

        print("Step 2: Creating multi-tab Excel file...")
        save_to_excel(data, output_file)
        print()

        print("=" * 60)
        print("[SUCCESS] EXCEL GENERATION COMPLETE")
        print("=" * 60)
        print(f"Output file: {output_file}")
        print()
        print("Tabs created:")
        print("  1. Metadata (Carrier, Validity info)")

        stats = data.get('statistics', {})
        if stats.get('MainCosts_sections', 0) > 0:
            print(f"  2. MainCosts ({stats.get('MainCosts_rows', 0)} pricing rows)")
        if stats.get('AddedRates_rows', 0) > 0:
            print(f"  3. AddedRates ({stats.get('AddedRates_rows', 0)} rows)")
        if stats.get('AdditionalCostsPart1_rows', 0) > 0:
            print(f"  4. AdditionalCostsPart1 ({stats.get('AdditionalCostsPart1_rows', 0)} rows)")
        if stats.get('CountryZoning_rows', 0) > 0:
            print(f"  5. CountryZoning ({stats.get('CountryZoning_rows', 0)} rows)")
        if stats.get('AdditionalZoning_rows', 0) > 0:
            print(f"  6. AdditionalZoning ({stats.get('AdditionalZoning_rows', 0)} rows)")
        if stats.get('ZoningMatrix_rows', 0) > 0:
            print(f"  7. ZoningMatrix ({stats.get('ZoningMatrix_rows', 0)} rows)")
        if stats.get('AdditionalCostsPart2_rows', 0) > 0:
            print(f"  8. AdditionalCostsPart2 ({stats.get('AdditionalCostsPart2_rows', 0)} rows)")
        acc_count = len(data.get('AdditionalCostsPart1', [])) + len(data.get('AdditionalCostsPart2', []))
        if acc_count > 0:
            print(f"  9. Accessorial Costs ({acc_count} rows)")
        print()

    except Exception as e:
        print()
        print("=" * 60)
        print("[FAILED] EXCEL GENERATION FAILED")
        print("=" * 60)
        print(f"Error: {e}")
        print()
        raise


# This block only runs when the script is executed directly.
# It does NOT run when this file is imported as a module by another script.
if __name__ == "__main__":
    main()




