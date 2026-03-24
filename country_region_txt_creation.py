"""
CountryZoning TXT Generator

HOW THIS FILE FITS INTO THE BIGGER PICTURE
-------------------------------------------
After create_table.py builds the Excel workbook, this script reads the
CountryZoning tab from that Excel file and produces a compact plain-text summary.

The TXT file lists each rate name (e.g. "Express Worldwide") followed by all the
country codes that belong to that rate, separated by commas.  It looks like this:

  Express Worldwide  DE, FR, IT, ES, NL, BE
  Economy Select     GB, IE, DK, SE, NO

This file is useful for quickly checking which countries are covered by each rate
without opening the full Excel workbook.

Output is saved to the same folder as the Excel file (or to a custom path).
"""

from pathlib import Path           # cross-platform file path handling
from collections import defaultdict  # used to build a dict of lists (rate name -> [countries])


def create_country_region_txt(
    excel_path: str = "output/DHL_Rate_Cards.xlsx",
    sheet_name: str = "CountryZoning",
    output_path: str | None = None,
) -> str:
    """
    Read the CountryZoning tab from an Excel workbook, group countries by rate name,
    and write a plain-text summary file.

    HOW IT WORKS:
      1. Open the Excel file and find the CountryZoning sheet.
      2. Read the header row to find which columns are "RateName" and "Country Code".
      3. Loop through all data rows, grouping country codes under their rate name.
         If a row's RateName cell is blank, the previous row's rate name is reused
         (forward-fill), because the Excel sheet may have merged cells for the rate name.
      4. Write one line per rate name: "RateName  code1, code2, code3, ..."
      5. Return the path of the created TXT file.

    Parameters:
      excel_path   – path to the Excel workbook to read (default: output/DHL_Rate_Cards.xlsx)
      sheet_name   – name of the sheet to read (default: "CountryZoning")
      output_path  – where to save the TXT file; if None, saves next to the Excel file

    Returns the path of the created TXT file as a string.
    """
    # openpyxl is the library used to read Excel files; it's imported here (not at the top)
    # so that the rest of the project still works even if openpyxl isn't installed
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    excel_path = Path(excel_path)
    print(f"[*] TXT Debug: excel_path={excel_path}")
    print(f"[*] TXT Debug: sheet_name={sheet_name}")

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Determine where to save the TXT file
    if output_path is None:
        # Default: save in the same folder as the Excel file
        output_dir = excel_path.parent
        output_path = output_dir / "CountryZoning_by_RateName.txt"
    else:
        output_path = Path(output_path)

    # Open the Excel file in read-only mode (faster; we only need to read, not write)
    # data_only=True means we get the cell values, not the formulas
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Check that the CountryZoning sheet exists in this workbook.
    # Some rate cards don't have country zoning data, so the sheet may be absent.
    if sheet_name not in wb.sheetnames:
        wb.close()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("", encoding="utf-8")   # write an empty file as a placeholder
        print(f"[WARN] Sheet '{sheet_name}' not found in {excel_path} (no CountryZoning data in this rate card). Wrote empty TXT: {output_path}")
        return str(output_path)

    print(f"[*] TXT Debug: workbook sheets={wb.sheetnames}")

    # Read all rows from the CountryZoning sheet into memory at once
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))   # each row is a tuple of cell values
    wb.close()
    print(f"[*] TXT Debug: total rows read (including header)={len(rows)}")

    if not rows:
        # The sheet exists but is completely empty
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("", encoding="utf-8")
        print("[WARN] TXT Debug: sheet is empty, wrote empty txt")
        return str(output_path)

    # -----------------------------------------------------------------------
    # Find the column positions for "RateName" and "Country Code".
    # We do this by reading the first row (the header row) and searching for
    # those exact column names.  This way the code still works even if the
    # columns are in a different order.
    # -----------------------------------------------------------------------
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    rate_name_col = None
    country_col = None
    for i, h in enumerate(headers):
        if h == "RateName":
            rate_name_col = i
        if h == "Country Code":
            country_col = i

    print(f"[*] TXT Debug: headers={headers}")
    print(f"[*] TXT Debug: RateName col index={rate_name_col}, Country Code col index={country_col}")

    if rate_name_col is None:
        raise ValueError("Column 'RateName' not found in CountryZoning")
    if country_col is None:
        raise ValueError("Column 'Country Code' not found in CountryZoning")

    # -----------------------------------------------------------------------
    # Loop through all data rows (skipping the header row) and group country
    # codes by rate name.
    #
    # FORWARD-FILL LOGIC:
    # In the Excel sheet, the RateName column may have blank cells for rows
    # that belong to the same rate as the row above (because the cells are
    # visually merged in Excel).  When we read the sheet, merged cells appear
    # as blank after the first row.  We handle this by remembering the last
    # non-blank rate name and reusing it for blank cells.
    #
    # Example:
    #   Row 2: RateName="Express Worldwide", Country Code="DE"
    #   Row 3: RateName="",                  Country Code="FR"   <- blank; reuse "Express Worldwide"
    #   Row 4: RateName="",                  Country Code="IT"   <- blank; reuse "Express Worldwide"
    #   Row 5: RateName="Economy Select",    Country Code="GB"   <- new rate name
    # -----------------------------------------------------------------------
    by_rate_name = defaultdict(list)   # maps rate name -> list of country codes
    current_rate = ""                  # the most recently seen non-blank rate name
    processed_rows = 0
    skipped_empty_country = 0

    for row in rows[1:]:   # rows[1:] skips the header row
        # Safely read the rate name and country code cells (guard against short rows)
        rn = row[rate_name_col] if rate_name_col < len(row) else None
        country = row[country_col] if country_col < len(row) else None

        # Update current_rate if this row has a non-blank rate name
        if rn is not None and str(rn).strip():
            current_rate = str(rn).strip()

        # Skip rows with no country code (e.g. blank rows between sections)
        if country is None or (isinstance(country, str) and not str(country).strip()):
            skipped_empty_country += 1
            continue

        # Add this country code to the list for the current rate name
        country_str = str(country).strip()
        if country_str:
            by_rate_name[current_rate].append(country_str)
            processed_rows += 1

    print(f"[*] TXT Debug: processed country rows={processed_rows}")
    print(f"[*] TXT Debug: skipped rows with empty Country Code={skipped_empty_country}")
    print(f"[*] TXT Debug: distinct RateName groups={len(by_rate_name)}")

    # -----------------------------------------------------------------------
    # Build the output lines and write the TXT file.
    #
    # Each line has the format:  RateName  code1, code2, code3, ...
    # Rate names are sorted alphabetically, with any blank rate name placed last.
    # -----------------------------------------------------------------------
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = []

    # Sort rate names alphabetically; the lambda puts blank names at the end
    for rate_name in sorted(by_rate_name.keys(), key=lambda x: (x == "", x)):
        countries = by_rate_name[rate_name]
        line = f"{rate_name}  {', '.join(countries)}"
        lines.append(line)

    print(f"[*] TXT Debug: output lines={len(lines)}")
    if lines:
        print(f"[*] TXT Debug: first line preview={lines[0][:200]}")
    else:
        print("[WARN] TXT Debug: no lines generated, output will be empty")

    # Write all lines to the TXT file, separated by newlines
    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[OK] TXT Debug: wrote file {output_path}")
    return str(output_path)


def main():
    """
    Entry point when this script is run directly from the command line.

    Looks for the Excel file at output/DHL_Rate_Cards.xlsx (relative to the
    script's own folder) and saves the TXT file in the same output/ folder.
    """
    # Resolve paths relative to the folder where this script lives,
    # so the script works regardless of where it is run from
    script_dir = Path(__file__).resolve().parent
    excel_path = script_dir / "output" / "DHL_Rate_Cards.xlsx"
    output_path = script_dir / "output" / "CountryZoning_by_RateName.txt"

    print("Creating CountryZoning TXT from DHL_Rate_Cards.xlsx...")
    out = create_country_region_txt(
        excel_path=str(excel_path),
        output_path=str(output_path),
    )
    print(f"Saved: {out}")


# Only run main() when this script is executed directly (e.g. python country_region_txt_creation.py).
# Does NOT run when imported as a module by pipeline_main.py.
if __name__ == "__main__":
    main()
