"""
Convert extracted JSON data to multi-tab Excel format for analysis.
Creates one tab per extracted field block (MainCosts, AddedRates, CountryZoning, etc.).

Input:  processing/extracted_data.json  (produced by the extraction pipeline)
Output: output/DHL_Rate_Cards.xlsx      (multi-tab workbook)

HOW THIS FILE FITS INTO THE BIGGER PICTURE
-------------------------------------------
Before this script runs, another script has already read a DHL rate-card PDF and
saved all the pricing data into a single JSON file (extracted_data.json).
This script's job is to take that JSON file and turn it into a nicely formatted
Excel workbook with one tab per data section, so analysts can open it directly.
"""

# --- Standard library imports ---
import difflib   # used for fuzzy text matching (finding the closest cost-type name)
import json      # used to read the extracted JSON file from disk
import os        # used to check file sizes and build file paths
import re        # used for pattern matching in text (e.g. finding "Zone 1" patterns)
from pathlib import Path   # used for cross-platform file path handling


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------

def load_extracted_data(filepath):
    """
    Open the extracted JSON file from disk and return its contents as a Python dictionary.
    If the file cannot be opened or is not valid JSON, an error is printed and the
    program stops immediately.
    """
    print(f"[*] Loading extracted data from: {filepath}")
    try:
        # Open the file in read mode with UTF-8 encoding (handles special characters)
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)   # parse the entire JSON file into a Python dict
        print(f"[OK] Data loaded successfully")
        return data
    except Exception as e:
        print(f"[ERROR] Failed to load data: {e}")
        raise   # re-raise so the calling code also knows something went wrong


# ---------------------------------------------------------------------------
# Metadata tab
# ---------------------------------------------------------------------------

def create_metadata_sheet(workbook, metadata):
    """
    Create the first tab in the Excel file called "Metadata".
    This tab shows basic information about the rate card document:
    who the client is, which carrier it belongs to, when it is valid, etc.
    It is a simple two-column table: column A = field name, column B = value.
    """
    print("[*] Creating Metadata tab...")
    
    # Import Excel styling tools (only loaded when this function actually runs)
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Add a new sheet called "Metadata" and insert it at position 0 (first tab)
    ws = workbook.create_sheet("Metadata", 0)
    
    # Define the blue header style (dark blue background, white bold text)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Helper: safely convert any value to a plain string.
    # - If the value is missing (None), return an empty string instead of crashing.
    # - If the value contains line breaks (e.g. carrier name on two lines), replace them with a space.
    def _str(v):
        s = "" if v is None else (v.replace("\n", " ") if isinstance(v, str) else str(v))
        return s

    # Build the table as a list of [field_name, value] pairs.
    # metadata is the dictionary that was read from the JSON file's "metadata" section.
    data = [
        ["Field", "Value"],                                          # header row
        ["Client", _str(metadata.get("client"))],                   # e.g. "Acme Corp"
        ["Carrier", _str(metadata.get("carrier"))],                 # e.g. "DHL Express France"
        ["Validity Date", _str(metadata.get("validity_date"))],     # e.g. "2024-01-01"
        ["FileName", _str(metadata.get("FileName"))],               # original PDF filename
        ["Extraction Date", _str(metadata.get("extraction_date"))], # when the PDF was processed
        ["Extraction Source", _str(metadata.get("extraction_source"))],  # tool/version used
    ]
    
    # Loop over every row and every cell in that row, and write it into the Excel sheet.
    # enumerate(data, 1) means row numbers start at 1 (Excel rows are 1-based).
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                # Only the very first row (the "Field / Value" header) gets the blue styling
                cell.fill = header_fill
                cell.font = header_font
            # All cells: allow text to wrap inside the cell and align to the top
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Set fixed column widths so the content is readable without manual resizing
    ws.column_dimensions['A'].width = 20   # field name column
    ws.column_dimensions['B'].width = 60   # value column (wider because values can be long)
    
    print(f"[OK] Metadata tab created")


# ---------------------------------------------------------------------------
# MainCosts – legacy flat pivot (zones as rows, weights as columns)
# ---------------------------------------------------------------------------

def pivot_main_costs(main_costs, metadata):
    """
    (Legacy / unused view) Convert the MainCosts pricing data into a simple flat table
    where each row = one delivery zone, and each column = one weight bracket.

    Example of what the output looks like:
        Zone    | 0.5 KG | 1 KG | 2 KG
        Zone 1  |  12.50 | 15.00| 18.00
        Zone 2  |  14.00 | 17.50| 21.00

    This is an older, simpler view.  The main view used today is build_matrix_main_costs().
    """
    rows = []   # will hold all the output rows we build

    # Pull the three identity fields that appear on every row
    client = (metadata.get('client') or '')
    carrier = (metadata.get('carrier') or '').replace('\n', ' ')  # remove any line breaks
    validity_date = (metadata.get('validity_date') or '')

    # Loop over each "rate card" block in the MainCosts list.
    # Each rate card covers one service type (e.g. "DHL EXPRESS WORLDWIDE EXPORT")
    # and one cost category (e.g. "Documents").
    for section_idx, rate_card in enumerate(main_costs, 1):
        service_type = rate_card.get('service_type') or ''
        cost_category = rate_card.get('cost_category', '')
        weight_unit = rate_card.get('weight_unit', 'KG')

        # zone_headers maps internal short keys (e.g. "Z1") to display names (e.g. "Zone 1")
        zone_headers = rate_card.get('zone_headers', {})

        # pricing is a list where each entry covers one weight breakpoint.
        # Example entry: { "weight": "0.5", "zone_prices": {"Z1": 12.50, "Z2": 14.00} }
        pricing = rate_card.get('pricing', [])
        
        # ---------------------------------------------------------------
        # Step 1: Reorganise the data from "weight-first" to "zone-first".
        #
        # The raw JSON groups prices by weight:
        #   weight 0.5 -> Zone1=12.50, Zone2=14.00
        #   weight 1.0 -> Zone1=15.00, Zone2=17.50
        #
        # We need it grouped by zone instead, so we can write one row per zone:
        #   Zone1 -> 0.5=12.50, 1.0=15.00
        #   Zone2 -> 0.5=14.00, 1.0=17.50
        # ---------------------------------------------------------------
        zone_price_matrix = {}   # zone_name -> { weight -> price }
        weights_set = set()      # collect all unique weight values seen

        for price_entry in pricing:
            weight = price_entry.get('weight', '')
            weights_set.add(weight)   # remember this weight so we can sort them later
            zone_prices = price_entry.get('zone_prices', {})

            for zone_key, price in zone_prices.items():
                # Translate the internal key (e.g. "Z1") into the display name (e.g. "Zone 1").
                # If no translation exists, use the key itself as a fallback.
                zone_name = zone_headers.get(zone_key, zone_key)
                if zone_name not in zone_price_matrix:
                    zone_price_matrix[zone_name] = {}   # create a new sub-dict for this zone
                zone_price_matrix[zone_name][weight] = price   # store the price
        
        # Sort the weight values numerically (0.5 before 1 before 2 etc.)
        # so the Excel columns appear in a logical order.
        try:
            weights_sorted = sorted(weights_set, key=lambda x: float(x))
        except:
            weights_sorted = sorted(weights_set)   # fallback: alphabetical sort
        
        # ---------------------------------------------------------------
        # Step 2: Build one output row per zone.
        # Each weight becomes its own column, named e.g. "0.5 KG".
        # ---------------------------------------------------------------
        for zone_name, weight_prices in zone_price_matrix.items():
            # Start the row with the fixed identity columns
            row = {
                'Client': client,
                'Carrier': carrier,
                'Validity Date': validity_date,
                'Section': section_idx,        # which rate card block this came from
                'Service Type': service_type,
                'Cost Category': cost_category,
                'Weight Unit': weight_unit,
                'Zone': zone_name
            }
            
            # Add one column per weight, e.g. "0.5 KG": 12.50
            # If this zone has no price for a particular weight, leave the cell empty ('')
            for weight in weights_sorted:
                col_name = f"{weight} {weight_unit}"   # e.g. "0.5 KG"
                row[col_name] = weight_prices.get(weight, '')
            
            rows.append(row)
    
    return rows


# ---------------------------------------------------------------------------
# Zone-name helpers
# ---------------------------------------------------------------------------

def _zone_has_letters(zone_name):
    """
    Check whether a zone name uses a letter identifier (e.g. "Zone A") rather than
    a number identifier (e.g. "Zone 1").

    WHY THIS MATTERS:
    - Numbered zones (Zone 1, Zone 2 …) directly represent a delivery region.
    - Letter zones (Zone A, Zone B …) are abstract codes that need to be looked up
      in the ZoningMatrix to find out which actual Origin/Destination countries they cover.
      So if a zone has letters, we know extra expansion work is needed later.

    Returns True for "Zone A", "Zone E", etc.
    Returns False for "Zone 1", "Zone 12", etc.
    """
    s = (zone_name or '').strip()
    if not s:
        return False   # empty string: treat as no letters

    # Remove the "Zone " prefix so we only examine the identifier part.
    # e.g. "Zone A" -> "A",  "Zone 1" -> "1"
    if s.upper().startswith('ZONE '):
        suffix = s[5:].strip()
    else:
        suffix = s

    # Return True if any character in the identifier is a letter
    return any(c.isalpha() for c in suffix)


# ---------------------------------------------------------------------------
# ZoningMatrix parsing and lane expansion
# ---------------------------------------------------------------------------

def parse_zoning_matrix(zoning_matrix):
    """
    Read the ZoningMatrix data and build a lookup table that answers the question:
    "For zone letter A in matrix X, which (origin zone, destination zone) pairs exist?"

    BACKGROUND – what is a ZoningMatrix?
    The ZoningMatrix is a grid that maps pairs of origin and destination zone numbers
    to a single letter (A, B, C …).  For example:
        Origin 1 -> Destination 3 -> letter "A"
        Origin 2 -> Destination 3 -> letter "A"
        Origin 1 -> Destination 5 -> letter "E"

    The MainCosts pricing table uses those letters as shorthand: instead of listing
    a price for every individual origin/destination pair, it lists one price per letter.
    This function reverses the matrix so we can later expand each letter back into
    all the concrete (origin, destination) pairs it represents.

    THE JSON STRUCTURE:
    The ZoningMatrix arrives as a flat list of rows.  Two types of rows alternate:
      - Header row: has 'MatrixName' filled in + DestinationZone1, DestinationZone2 …
                    whose values are the destination zone numbers (1, 2, 3 …)
      - Data row:   has 'OriginZone' filled in + DestinationZone1, DestinationZone2 …
                    whose values are the zone letters (A, B, E …)

    WHAT THIS FUNCTION RETURNS:
    A dictionary where:
      key   = (matrix_name, zone_letter)   e.g. ("DHL EXPRESS WW ZONE MATRIX", "A")
      value = list of (origin_zone, destination_zone) pairs  e.g. [("1", "3"), ("2", "3")]
    """
    import re
    result = {}                    # the lookup table we are building
    dest_cols = None               # ordered list of "DestinationZone1", "DestinationZone2" … keys
    header_dest_nums = None        # the actual destination zone numbers read from the header row
    current_matrix_name = None     # name of the matrix block we are currently inside

    for row in zoning_matrix or []:
        matrix_name = (row.get('MatrixName') or '').strip()
        origin_zone = (row.get('OriginZone') or '').strip()

        if matrix_name:
            # ---------------------------------------------------------------
            # This is a HEADER ROW – it starts a new matrix block.
            # Example: MatrixName="DHL EXPRESS WW ZONE MATRIX",
            #          DestinationZone1="1", DestinationZone2="2", DestinationZone3="3"
            # ---------------------------------------------------------------
            current_matrix_name = matrix_name

            # Find all keys that look like "DestinationZone1", "DestinationZone2" etc.
            # and sort them numerically so column order is preserved.
            dest_keys = sorted(
                [k for k in row if re.match(r'^DestinationZone\d+$', k)],
                key=lambda k: int(re.search(r'\d+', k).group())
            )
            dest_cols = dest_keys

            # Read the actual destination zone numbers from the header cells.
            # e.g. DestinationZone1 -> "1", DestinationZone2 -> "2"
            header_dest_nums = [str(row.get(k, '')).strip() for k in dest_cols]
            continue   # move on to the next row (this header row has no prices)

        if current_matrix_name and origin_zone and dest_cols:
            # ---------------------------------------------------------------
            # This is a DATA ROW – it belongs to the current matrix block.
            # Example: OriginZone="1",
            #          DestinationZone1="A", DestinationZone2="A", DestinationZone3="E"
            # This means: origin 1 -> destination 1 = letter A
            #             origin 1 -> destination 2 = letter A
            #             origin 1 -> destination 3 = letter E
            # ---------------------------------------------------------------
            for col_idx, dest_key in enumerate(dest_cols):
                if col_idx >= len(header_dest_nums):
                    continue   # safety check: don't go past the number of header columns
                dest_zone_num = header_dest_nums[col_idx]   # e.g. "3"
                if not dest_zone_num:
                    continue   # skip if the header had no zone number for this column
                cell_letter = (row.get(dest_key) or '').strip()   # e.g. "A"
                if not cell_letter:
                    continue   # skip empty cells (no zone letter assigned)

                # Build the lookup key: (matrix_name, letter)
                key = (current_matrix_name, cell_letter.upper())
                if key not in result:
                    result[key] = []   # create a new list for this letter if first time seen
                # Record that this (origin, destination) pair maps to this letter
                result[key].append((origin_zone, dest_zone_num))

    return result


def _matrix_zone_to_letter(matrix_zone):
    """
    Extract just the letter part from a zone name like "Zone E" -> "E".
    This is needed because the lookup table is keyed by the letter alone, not the full name.
    If the input is already just a letter (no "Zone " prefix), it is returned as-is in uppercase.
    """
    s = (matrix_zone or '').strip()
    if not s:
        return ''
    if s.upper().startswith('ZONE '):
        return s[5:].strip().upper()   # remove "Zone " and return the rest in uppercase
    return s.upper()


def _main_words(text):
    """
    Split a text string into its meaningful words (all uppercase), ignoring the
    generic words "ZONE" and "MATRIX" which appear in almost every matrix name
    and would cause false matches.

    Example: "DHL EXPRESS THIRD COUNTRY ZONE MATRIX" -> {"DHL", "EXPRESS", "THIRD", "COUNTRY"}
    """
    if not text:
        return set()
    words = set((text or '').upper().split())
    words.discard('ZONE')     # too generic to be useful for matching
    words.discard('MATRIX')   # too generic to be useful for matching
    return words


def _find_matrix_for_service(zoning_lookup, service):
    """
    Given a service type name (e.g. "DHL EXPRESS THIRD COUNTRY"), find which matrix
    in the zoning_lookup corresponds to it.

    WHY THIS IS NEEDED:
    The service names in MainCosts and the matrix names in ZoningMatrix are written
    slightly differently.  For example:
      - Service:  "DHL EXPRESS THIRD COUNTRY"
      - Matrix:   "DHL EXPRESS THIRD COUNTRY ZONE MATRIX"
    We need to match them up despite these differences.

    MATCHING STRATEGY (tries each approach in order, returns the first match found):
      1. Direct substring: does the service name appear inside the matrix name, or vice versa?
      2. Strip " ZONE MATRIX" from the matrix name, then try substring again.
      3. Word-level match: do all meaningful words from the matrix name appear in the service?
         e.g. {"DHL", "EXPRESS", "THIRD", "COUNTRY"} are all present in "DHL EXPRESS THIRD COUNTRY"

    Returns the matching matrix name, or None if no match is found.
    """
    service = (service or '').strip()
    if not service:
        return None
    service_words = _main_words(service)

    # Get all unique matrix names from the lookup (ignoring the zone letter part of each key)
    matrix_names = {mn for (mn, _) in zoning_lookup}

    # --- Attempt 1: direct substring match ---
    for mn in matrix_names:
        if service in mn or mn in service:
            return mn   # found a match, return immediately

    # --- Attempt 2: strip the " ZONE MATRIX" boilerplate and try again ---
    for mn in matrix_names:
        normalized = mn.replace(' ZONE MATRIX', '').strip()
        if service in normalized or normalized in service:
            return mn

    # --- Attempt 3: all meaningful words from the matrix name must be in the service ---
    # This handles cases where word order differs or extra words are present
    for mn in matrix_names:
        matrix_words = _main_words(mn.replace(' ZONE MATRIX', ''))
        # "<=" on sets means "is a subset of": all matrix words appear in service words
        if matrix_words and matrix_words <= service_words:
            return mn

    return None   # no match found in any of the three attempts


def expand_main_costs_lanes_by_zoning(matrix_rows, zoning_matrix):
    """
    Replace abstract letter-zone rows with real Origin/Destination rows.

    PROBLEM THIS SOLVES:
    After build_matrix_main_costs() runs, some lanes have a "Matrix zone" value
    like "Zone A" instead of real origin/destination countries.  "Zone A" is just
    a code that means "all the origin/destination pairs that belong to group A".
    This function looks up those pairs and creates one concrete row per pair.

    EXAMPLE:
    Before expansion:
        Lane | Origin | Destination | Service          | Matrix zone | Price
        1    |        |             | DHL EXPRESS WW   | Zone A      | 12.50

    After expansion (if Zone A covers origin 1->dest 3 and origin 2->dest 3):
        Lane | Origin | Destination | Service          | Matrix zone | Price
        1    | Zone 1 | Zone 3      | DHL EXPRESS WW   | Zone A      | 12.50
        2    | Zone 2 | Zone 3      | DHL EXPRESS WW   | Zone A      | 12.50

    Rows that already have numeric zones (no Matrix zone value) are left unchanged.
    After all expansion is done, Lane numbers are reassigned from 1 upward.
    """
    if not matrix_rows:
        return matrix_rows   # nothing to do if there are no rows

    # Build the full (matrix_name, zone_letter) -> [(origin, dest), ...] lookup
    # by reading the ZoningMatrix data once.  This is the "reverse lookup" table.
    zoning_lookup = parse_zoning_matrix(zoning_matrix)
    if not zoning_lookup:
        return matrix_rows   # no zoning data available; return rows unchanged

    expanded = []   # will hold the final list of rows after expansion

    for row in matrix_rows:
        matrix_zone = (row.get('Matrix zone') or '').strip()
        service = (row.get('Service') or '').strip()

        # If this row has no Matrix zone value, it already has real Origin/Destination.
        # Just copy it through to the output without any changes.
        if not matrix_zone:
            expanded.append(row)
            continue

        # Extract the letter from e.g. "Zone A" -> "A"
        zone_letter = _matrix_zone_to_letter(matrix_zone)
        if not zone_letter:
            # Could not parse a letter; keep the row as-is
            expanded.append(row)
            continue

        # Find which matrix in the lookup table matches this service type
        matrix_name = _find_matrix_for_service(zoning_lookup, service)
        if not matrix_name:
            # No matching matrix found; keep the row as-is
            expanded.append(row)
            continue

        # Look up all (origin_zone, destination_zone) pairs for this letter
        key = (matrix_name, zone_letter)
        pairs = zoning_lookup.get(key, [])
        if not pairs:
            # This letter has no pairs in the matrix; keep the row as-is
            expanded.append(row)
            continue

        # For each (origin, destination) pair, create a separate copy of the row
        # with the Origin and Destination fields filled in with the real zone numbers.
        for origin_zone, dest_zone in pairs:
            new_row = row.copy()   # copy all existing fields (prices, service name, etc.)
            new_row['Origin'] = f"Zone {origin_zone}" if origin_zone else ''
            new_row['Destination'] = f"Zone {dest_zone}" if dest_zone else ''
            expanded.append(new_row)

    # After expansion, the original Lane # values are no longer sequential.
    # Reassign them from 1 upward so the numbering is clean and continuous.
    for lane, row in enumerate(expanded, 1):
        row['Lane #'] = lane

    return expanded


# ---------------------------------------------------------------------------
# Carrier country helper
# ---------------------------------------------------------------------------

def global_country(metadata):
    """
    Extract the country name from the carrier string in the metadata.

    DHL carrier names follow the pattern "DHL DHL Express <Country>",
    e.g. "DHL DHL Express France" -> "France".
    The last word is always the country.

    This country name is used to fill in the Origin or Destination column for:
    - Domestic lanes (both Origin and Destination = carrier's country)
    - Non-zoned export lanes (Destination = carrier's country)
    - Non-zoned import lanes (Origin = carrier's country)
    """
    carrier = (metadata.get('carrier') or '').replace('\n', ' ').strip()
    parts = carrier.split()          # split into individual words
    return parts[-1] if parts else ''   # return the last word, or '' if carrier is empty


# ---------------------------------------------------------------------------
# MainCosts – matrix (lane) view builder
# ---------------------------------------------------------------------------

def _zone_sort_key(zone_name):
    """
    Generate a sort key for a zone name so that zones appear in a sensible order:
    numeric zones first (Zone 1, Zone 2, Zone 10 …) then letter zones (Zone A, Zone B …).

    Without this, alphabetical sorting would give: Zone 1, Zone 10, Zone 2 (wrong).
    With this, we get: Zone 1, Zone 2, Zone 10, Zone A (correct).

    Returns a tuple (group, value) where:
      group=0 means numeric zone (sorted by number)
      group=1 means letter zone (sorted after all numeric zones)
    """
    s = (zone_name or '').strip()
    if not s:
        return (1, 0)   # empty zone name: put at the end
    if s.upper().startswith('ZONE '):
        suffix = s[5:].strip()   # e.g. "Zone 3" -> "3"
    else:
        suffix = s
    try:
        # If the suffix is a number, sort numerically within group 0
        return (0, float(suffix))
    except (ValueError, TypeError):
        # If the suffix is a letter (or anything non-numeric), put it in group 1
        return (1, 0 if _zone_has_letters(zone_name) else 1)


def build_matrix_main_costs(main_costs, metadata):
    """
    Build the main pricing table (called the "Matrix view") for the MainCosts Excel tab.

    WHAT THE OUTPUT LOOKS LIKE:
    Each output row = one "lane" = one unique combination of service type + zone.
    All cost categories (Envelope, Documents, Parcels …) for the same lane are
    combined into a single row, with prices stored as separate columns per weight.

    Example output row:
        Lane# | Origin | Destination | Service              | Matrix zone | Envelope 0.5KG | Envelope 1KG | Documents 0.5KG …
        1     | France | Zone 1      | DHL EXPRESS EXPORT   |             | 12.50          | 15.00        | 10.00 …

    Returns two things:
      rows           – the list of lane rows described above
      category_specs – a description of each cost-category column group,
                       used by write_matrix_sheet() to draw the header
    """
    # =======================================================================
    # PASS 1 – Figure out what columns the header needs.
    #
    # Before we can write any data, we need to know ALL cost categories and
    # ALL weight breakpoints that exist anywhere in the data, so the header
    # can be drawn correctly.  We scan every rate card once to collect this.
    # =======================================================================
    category_specs = []   # will hold: [(category_name, weight_unit, [0.5, 1, 2, …]), …]
    seen_categories = {}  # tracks which categories we have already added (for deduplication)

    for rate_card in main_costs:
        cost_category = rate_card.get('cost_category') or ''   # e.g. "Documents"
        weight_unit = rate_card.get('weight_unit') or 'KG'     # e.g. "KG"
        pricing = rate_card.get('pricing', [])

        # Collect all weight values mentioned in this rate card's pricing list
        weights_set = set()
        for pe in pricing:
            w = pe.get('weight', '')
            if w:
                weights_set.add(w)

        # Sort the weights numerically (0.5, 1, 2 …) for a clean column order
        try:
            weights_sorted = sorted(weights_set, key=lambda x: float(x))
        except (ValueError, TypeError):
            weights_sorted = sorted(weights_set)

        if cost_category not in seen_categories:
            # First time we see this category: add it to the list
            seen_categories[cost_category] = (weight_unit, weights_sorted)
            category_specs.append((cost_category, weight_unit, weights_sorted))
        else:
            # We have seen this category before (in a different rate card).
            # Merge the weight sets so we capture all possible breakpoints.
            existing_unit, existing_weights = seen_categories[cost_category]
            merged = set(existing_weights) | set(weights_sorted)   # union of both sets
            try:
                merged_sorted = sorted(merged, key=lambda x: float(x))
            except (ValueError, TypeError):
                merged_sorted = sorted(merged)
            seen_categories[cost_category] = (existing_unit, merged_sorted)
            # Update the spec we already stored so it reflects the merged weights
            for i, spec in enumerate(category_specs):
                if spec[0] == cost_category:
                    category_specs[i] = (cost_category, existing_unit, merged_sorted)
                    break

    # =======================================================================
    # PASS 2 – Build one row per lane (service + zone combination).
    #
    # We go through every rate card again.  For each zone in that rate card,
    # we either create a new lane row or add prices to an existing one.
    # Prices are stored with a composite key (category_name, weight) so they
    # can be placed in the correct column later.
    # =======================================================================
    lane_rows = {}   # (service_type, zone_name) -> row dict

    for rate_card in main_costs:
        service_type = (rate_card.get('service_type') or '').strip()   # e.g. "DHL EXPRESS WORLDWIDE EXPORT"
        cost_category = rate_card.get('cost_category') or ''           # e.g. "Documents"
        zone_headers = rate_card.get('zone_headers', {})               # e.g. {"Z1": "Zone 1"}
        pricing = rate_card.get('pricing', [])

        service_lower = service_type.lower()
        is_import = 'import' in service_lower   # True if this is an import service
        is_export = 'export' in service_lower   # True if this is an export service

        # Reorganise the pricing list from weight-first to zone-first
        # (same inversion as in pivot_main_costs)
        zone_price_matrix = {}
        for price_entry in pricing:
            weight = price_entry.get('weight', '')
            zone_prices = price_entry.get('zone_prices', {})
            for zone_key, price in zone_prices.items():
                zone_name = zone_headers.get(zone_key, zone_key)   # translate key to display name
                if zone_name not in zone_price_matrix:
                    zone_price_matrix[zone_name] = {}
                zone_price_matrix[zone_name][weight] = price

        for zone_name, weight_prices in zone_price_matrix.items():
            key = (service_type, zone_name)   # unique identifier for this lane

            if key not in lane_rows:
                # First time we encounter this (service, zone) combination:
                # create the base row with the identity fields.
                #
                # For import services, the zone is the Origin (where the parcel comes from).
                # For export services, the zone is the Destination (where it goes to).
                # Letter zones (A, B …) will be replaced later by expand_main_costs_lanes_by_zoning().
                origin = zone_name if is_import else ''
                destination = zone_name if is_export else ''
                matrix_zone = zone_name if _zone_has_letters(zone_name) else ''
                lane_rows[key] = {
                    'Origin': origin,
                    'Destination': destination,
                    'Service': service_type,
                    'Matrix zone': matrix_zone,   # filled only for letter-based zones
                }

            row = lane_rows[key]
            # Add the prices from this cost category into the existing lane row.
            # The key (cost_category, weight) acts like a column address, e.g.
            # ("Documents", "0.5") -> 10.00
            for weight, price in weight_prices.items():
                row[(cost_category, weight)] = price

    # Get the carrier's country (e.g. "France") to fill blank Origin/Destination fields
    carrier_last = global_country(metadata)

    # Sort the lanes: first by service name (alphabetical), then by zone (numeric before letter)
    sorted_keys = sorted(lane_rows.keys(), key=lambda k: (k[0], _zone_sort_key(k[1])))

    rows = []
    for lane, key in enumerate(sorted_keys, 1):
        row = lane_rows[key].copy()
        row['Lane #'] = lane   # assign a sequential lane number

        service = (row.get('Service') or '').strip()
        matrix_zone = (row.get('Matrix zone') or '').strip()

        if service == 'DHL EXPRESS DOMESTIC':
            # Domestic service: the parcel stays within the carrier's country,
            # so both Origin and Destination are the carrier's country.
            if carrier_last:
                row['Origin'] = carrier_last
                row['Destination'] = carrier_last
        elif not matrix_zone:
            # Non-zoned lane (no letter zone): the carrier's country fills whichever
            # of Origin or Destination is still empty.
            # e.g. an export lane already has Destination=zone; Origin gets the country.
            if carrier_last:
                if not (row.get('Origin') or '').strip():
                    row['Origin'] = carrier_last
                if not (row.get('Destination') or '').strip():
                    row['Destination'] = carrier_last

        rows.append(row)

    return rows, category_specs


# ---------------------------------------------------------------------------
# CountryZoning helpers
# ---------------------------------------------------------------------------

def _transform_rate_name_to_short(rate_name):
    """
    Convert a long rate card name into a short, underscore-separated code.

    WHY THIS IS NEEDED:
    In the CountryZoning tab, only the first row of each zone block has a full
    RateName (e.g. "DHL EXPRESS WORLDWIDE EXPORT ZONING").  The rows that follow
    (one per country) have an empty RateName.  We fill those empty cells with a
    short version of the name plus the zone, e.g. "WW_EXP_ZONE_Zone 1".

    TRANSFORMATION RULES (applied in order):
      "DHL EXPRESS"    -> removed entirely (it's on every name, adds no value)
      "THIRD COUNTRY"  -> "3RD_COUNTRY"
      "INTERNATIONAL"  -> "WW"  (worldwide)
      "IMPORT"         -> "IMP"
      "EXPORT"         -> "EXP"
      "ZONING"         -> "ZONE"

    The surviving tokens are then assembled in a fixed order so the result is
    always consistent regardless of the original word order:
      e.g. "DHL EXPRESS WORLDWIDE EXPORT ZONING" -> "WW_EXP_ZONE"
      e.g. "DHL EXPRESS THIRD COUNTRY IMPORT"    -> "3RD_COUNTRY_IMP"
    """
    if not rate_name or not isinstance(rate_name, str):
        return ''
    s = rate_name.upper().strip()   # work in uppercase to make matching case-insensitive

    # Apply each replacement in sequence.
    # Spaces around the replacements ensure the tokens are separated cleanly.
    s = s.replace('DHL EXPRESS', ' ')
    s = s.replace('THIRD COUNTRY', ' 3RD_COUNTRY ')
    s = s.replace('INTERNATIONAL', ' WW ')
    s = s.replace('IMPORT', ' IMP ')
    s = s.replace('EXPORT', ' EXP ')
    s = s.replace('ZONING', ' ZONE ')

    # Now scan for each known token in a fixed canonical order.
    # Only include tokens that actually appear in the transformed string.
    tokens = []
    for token in ('WW', '3RD_COUNTRY', 'DOMESTIC', 'ECONOMY', 'EXP', 'IMP', 'ZONE'):
        if token in s and token not in tokens:
            tokens.append(token)

    # Join the collected tokens with underscores, e.g. ["WW", "EXP", "ZONE"] -> "WW_EXP_ZONE"
    return '_'.join(tokens) if tokens else ''


def _fill_country_zoning_rate_names(rows):
    """
    Fill in the empty RateName cells in the CountryZoning rows.

    THE PROBLEM:
    In the source JSON, the CountryZoning data looks like this:
        Row 1: RateName="DHL EXPRESS WW EXPORT ZONING", Zone="Zone 1", Country="France"
        Row 2: RateName="",                             Zone="Zone 1", Country="Germany"
        Row 3: RateName="",                             Zone="Zone 1", Country="Spain"
        Row 4: RateName="DHL EXPRESS WW EXPORT ZONING", Zone="Zone 2", Country="France"
        ...

    Only the first row of each block has a RateName.  The rest are empty.
    We need to fill them in so every row has a meaningful RateName.

    THE SOLUTION:
    Walk through all rows in order, remembering the last non-empty RateName seen.
    When we find a row with an empty RateName, build a short name from the last
    remembered name plus the current Zone value.

    Example result for rows 2 and 3 above:
        RateName = "WW_EXP_ZONE_Zone 1"
    """
    last_rate_name = ''   # stores the most recently seen non-empty RateName

    for row in rows:
        rate_name = row.get('RateName') or ''
        zone = row.get('Zone') or ''

        if rate_name:
            # This row has a RateName – remember it for the rows that follow
            last_rate_name = rate_name

        if not rate_name and last_rate_name and zone:
            # This row has no RateName but we know the last one and the current zone.
            # Build a short name: convert the long name to a prefix, then append the zone.
            # e.g. "DHL EXPRESS WW EXPORT ZONING" -> "WW_EXP_ZONE" -> "WW_EXP_ZONE_Zone 1"
            prefix = _transform_rate_name_to_short(last_rate_name)
            if prefix:
                row['RateName'] = f"{prefix}_{zone}"


# ---------------------------------------------------------------------------
# Country code lookup
# ---------------------------------------------------------------------------

def _load_country_codes(codes_path=None):
    """
    Load the country-name-to-ISO-code dictionary from a plain text file.

    FILE FORMAT (one country per line, tab-separated):
        France    FR
        Germany   DE
        China     CN,CHN

    If a country has multiple codes separated by commas, only the first is used.

    The file is looked for in two locations (in order):
      1. input/dhl_country_codes.txt   (next to this script)
      2. addition/dhl_country_codes.txt

    Returns a dictionary like: {"France": "FR", "Germany": "DE", "China": "CN"}
    Returns an empty dict {} if the file is not found.
    """
    if codes_path is None:
        # Build the default path relative to the location of this script file
        base = Path(__file__).resolve().parent
        codes_path = base / "input" / "dhl_country_codes.txt"
        if not codes_path.exists():
            codes_path = base / "addition" / "dhl_country_codes.txt"   # try fallback location
    print(f"[*] CountryCode Debug: trying codes file: {codes_path}")
    codes_path = Path(codes_path)
    if not codes_path.exists():
        print(f"[WARN] CountryCode Debug: codes file not found: {codes_path}")
        return {}   # return empty dict; country codes will be blank in the output

    name_to_code = {}   # the dictionary we are building

    for line in codes_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or "\t" not in line:
            continue   # skip blank lines and lines without a tab separator

        # Split on the first tab only.
        # Everything before the tab = country name; everything after = code(s)
        name, code = line.split("\t", 1)
        name = name.strip()
        code = code.strip()

        # If the code field contains multiple codes (e.g. "CN,CHN"), use only the first
        if "," in code:
            code = code.split(",")[0].strip()

        if name:
            name_to_code[name] = code   # store the mapping

    print(f"[OK] CountryCode Debug: loaded mappings: {len(name_to_code)}")
    if name_to_code:
        sample_items = list(name_to_code.items())[:5]
        print(f"[*] CountryCode Debug: sample mappings: {sample_items}")
    return name_to_code


def _country_to_code(country, name_to_code):
    """
    Look up the ISO country code for a given country name string.

    The country names in the rate card data are not always written exactly the same
    way as in the reference file.  This function tries several variations to find a match.

    LOOKUP ATTEMPTS (in order, returns the first match found):
      1. Exact match as-is                   e.g. "France" -> "FR"
      2. Uppercase version                   e.g. "france" -> "FRANCE" -> "FR"
      3. Common name normalizations:
           "Republic Of" <-> "Rep. Of"       e.g. "Republic Of Korea" -> "Rep. Of Korea"
           " And " <-> " & "                 e.g. "Bosnia And Herzegovina" -> "Bosnia & Herzegovina"
           Strip ", Peoples Republic" etc.   e.g. "China, Peoples Republic" -> "China"
      4. Embedded code fallback:
           If the input is "Afghanistan (AF)", extract "AF" as a last resort.

    Returns the 2-letter (or 3-letter) code string, or '' if nothing matched.
    """
    if not country:
        return ''
    s = str(country).strip()
    if not s:
        return ''

    # Check if the country string already contains an ISO code in parentheses,
    # e.g. "Afghanistan (AF)".  Save the code as a fallback in case name lookup fails.
    paren_code = ''
    m = re.match(r'^(.*?)\s*\(([A-Za-z]{2,3})\)\s*$', s)
    if m:
        s = m.group(1).strip()        # remove the "(AF)" part; keep just "Afghanistan"
        paren_code = m.group(2).upper()   # save "AF" as fallback

    # Attempt 1: exact match
    code = name_to_code.get(s)
    if code is not None:
        return code

    # Attempt 2: uppercase exact match (handles case differences)
    code = name_to_code.get(s.upper())
    if code is not None:
        return code

    # Attempt 3: build a list of normalised variants and try each one
    variants = []
    # Normalise "Republic Of" -> "Rep. Of" (common abbreviation difference)
    n = s.replace("Republic Of", "Rep. Of").replace("Republic of", "Rep. Of")
    n = n.replace(", Republic", ", Rep.").replace(" Republic", " Rep.")
    variants.append(n)
    variants.append(n.replace(" And ", " & "))   # "Bosnia And Herzegovina" -> "Bosnia & Herzegovina"
    variants.append(n.replace(" & ", " And "))   # reverse: "Bosnia & Herzegovina" -> "Bosnia And Herzegovina"

    # Strip ", Peoples Republic" suffixes to get the base country name
    # e.g. "China, Peoples Republic" -> "China"
    for suffix in (", Peoples Republic", ", People's Republic", ", Peoples Rep.", ", People's Rep.",
                   " Peoples Republic", " People's Republic"):
        if n.endswith(suffix) or suffix in n:
            base = n.replace(suffix, "").strip().strip(",").strip()
            if base:
                variants.append(base)

    # Try each variant (both as-is and uppercase)
    for v in variants:
        if not v:
            continue
        code = name_to_code.get(v)
        if code is not None:
            return code
        code = name_to_code.get(v.upper())
        if code is not None:
            return code

    # Attempt 4: use the embedded parenthetical code as a last resort
    if paren_code:
        return paren_code

    return ''   # no match found at all


def _fill_country_zoning_country_codes(rows, name_to_code):
    """
    Add a 'Country Code' column to every CountryZoning row by looking up
    the value in the 'Country' column against the name_to_code dictionary.

    After this runs, each row will have a new 'Country Code' field, e.g. "FR" for France.
    Rows where the country name could not be matched will have an empty 'Country Code'.

    At the end, a summary is printed showing how many countries were matched vs missed,
    and a sample of up to 20 unmatched country names (to help diagnose data issues).
    """
    matched = 0          # counter: how many rows got a code
    missing = 0          # counter: how many rows had a country name but no code was found
    missing_countries = []   # sample list of unmatched names for the warning log

    for row in rows:
        country = row.get('Country') or ''
        code = _country_to_code(country, name_to_code)   # look up the code
        row['Country Code'] = code   # write the result back into the row

        if country and code:
            matched += 1   # successfully resolved
        elif country and not code:
            missing += 1   # had a name but couldn't find a code
            # Collect up to 20 unresolved names so we can log them as a warning
            if len(missing_countries) < 20:
                missing_countries.append(str(country))

    print(f"[*] CountryCode Debug: rows with country matched={matched}, missing={missing}")
    if missing_countries:
        print(f"[WARN] CountryCode Debug: sample missing countries: {missing_countries}")


# ---------------------------------------------------------------------------
# Generic array flattener (CountryZoning, AdditionalZoning, ZoningMatrix, …)
# ---------------------------------------------------------------------------

def flatten_array_data(array_data, metadata, field_name):
    """
    Convert a JSON array (list of objects) into a list of row dictionaries
    ready to be written to an Excel sheet.

    WHAT IT DOES:
    Each item in the JSON array becomes one row.  Before the item's own fields,
    three common identity columns are prepended to every row:
      - Client       (who the rate card belongs to)
      - Carrier      (which carrier, e.g. DHL Express France)
      - Validity Date (when the rates are valid from)

    SPECIAL HANDLING FOR CountryZoning:
    The CountryZoning data needs two extra enrichment steps that other arrays don't need:
      1. Forward-fill empty RateName cells (see _fill_country_zoning_rate_names)
      2. Add a Country Code column by looking up each country name (see _fill_country_zoning_country_codes)

    All other arrays (AdditionalZoning, ZoningMatrix, etc.) are passed through as-is
    with just the three identity columns prepended.
    """
    rows = []

    # Read the three identity values once (they are the same for every row)
    client = (metadata.get('client') or '')
    carrier = (metadata.get('carrier') or '').replace('\n', ' ')   # remove line breaks
    validity_date = (metadata.get('validity_date') or '')

    for item in array_data:
        # Create a new row dict starting with the three identity columns
        row = {
            'Client': client,
            'Carrier': carrier,
            'Validity Date': validity_date
        }
        # Merge all fields from the JSON item into the row.
        # If the item happens to have a key named "Client" or "Carrier", it would
        # overwrite the identity value – but in practice this does not happen.
        row.update(item)
        rows.append(row)
    
    if field_name == 'CountryZoning':
        # Step 1: fill in the empty RateName cells using the forward-fill logic
        _fill_country_zoning_rate_names(rows)
        # Step 2: load the country-name-to-code dictionary and add a Country Code column
        name_to_code = _load_country_codes()
        _fill_country_zoning_country_codes(rows, name_to_code)
    
    return rows


# ---------------------------------------------------------------------------
# AddedRates pivot
# ---------------------------------------------------------------------------

def _is_added_rates_header_row(item):
    """
    Decide whether a single AddedRates JSON item is a "header row" or a "data row".

    BACKGROUND:
    The AddedRates JSON mixes two types of rows together in one flat list:
      - Header rows: contain the zone names (e.g. "Zone 1", "Zone 2") in the Zone1, Zone2 … fields.
                     They also carry the table name and page reference.
      - Data rows:   contain actual weight ranges and prices (e.g. WeightFrom=0.5, Zone1=12.50).

    We detect header rows by checking:
      - WeightFrom == "From"  (the literal word "From" signals a header, not a weight value)
      - OR Zone1 value starts with "Zone"  (the cell contains a zone label, not a price)

    Returns True if this item is a header row, False if it is a data row.
    """
    weight_from = item.get('WeightFrom', '')
    zone1_val = item.get('Zone1', '')
    if weight_from == 'From' or (str(zone1_val).strip().startswith('Zone')):
        return True
    return False


def pivot_added_rates(added_rates, metadata):
    """
    Convert the AddedRates JSON list into a clean flat table for Excel.

    THE CHALLENGE:
    The source JSON for AddedRates looks like this (simplified):
        { WeightFrom:"From", WeightTo:"To", Zone1:"Zone 1", Zone2:"Zone 2", TableName:"Fuel Surcharge" }  <- header
        { WeightFrom:"0",    WeightTo:"0.5", Zone1:"12.50", Zone2:"14.00" }                              <- data
        { WeightFrom:"0.5",  WeightTo:"1",   Zone1:"15.00", Zone2:"17.50" }                              <- data
        { WeightFrom:"From", WeightTo:"To", Zone1:"Zone 1", Zone2:"Zone 2", TableName:"Remote Area" }    <- header (new table)
        ...

    Header rows tell us what the zone columns are called.
    Data rows contain the actual weight ranges and prices.

    WHAT THIS FUNCTION PRODUCES:
    Every JSON item becomes one output row.  For data rows, the Zone1, Zone2 … values
    are written under the human-readable column names taken from the most recent header row.
    Page Stopper and Table Name are only filled on header rows (they are blank on data rows).

    Example output:
        Client | Carrier | Validity Date | Page Stopper | Table Name     | Weight From | Weight To | Zone 1 | Zone 2
               |         |               | p.5          | Fuel Surcharge | From        | To        | Zone 1 | Zone 2   <- header row
               |         |               |              |                | 0           | 0.5       | 12.50  | 14.00    <- data row
               |         |               |              |                | 0.5         | 1         | 15.00  | 17.50    <- data row
    """
    rows = []
    client = (metadata.get('client') or '')
    carrier = (metadata.get('carrier') or '').replace('\n', ' ')
    validity_date = (metadata.get('validity_date') or '')

    # zone_column_names holds the current mapping from JSON key to display label.
    # e.g. [("Zone1", "Zone 1"), ("Zone2", "Zone 2")]
    # It is rebuilt every time we encounter a new header row.
    zone_column_names = []

    for item in added_rates:
        is_header = _is_added_rates_header_row(item)

        if is_header:
            # This is a header row.  The Zone1, Zone2 … cells contain the display labels
            # (e.g. "Zone 1", "Zone 2") that should be used as column names for all
            # data rows that follow until the next header row.
            zone_column_names = []
            zone_keys = [k for k in item.keys() if k.startswith('Zone')]

            # Sort the zone keys numerically: Zone1, Zone2, Zone3 … (not Zone1, Zone10, Zone2)
            def zone_sort_key(k):
                suffix = k[4:]   # the number part after "Zone", e.g. "1" from "Zone1"
                try:
                    return int(suffix)
                except ValueError:
                    return 0   # non-numeric suffix: put first as fallback

            for k in sorted(zone_keys, key=zone_sort_key):
                # Build a pair: (json_key, display_label)
                # e.g. ("Zone1", "Zone 1")  or  ("Zone2", "Zone 2")
                zone_column_names.append((k, str(item.get(k, k)).strip() or k))

        # Build the output row for this item (whether header or data)
        weight_from = item.get('WeightFrom', '')
        weight_to = item.get('WeightTo', '')
        row = {
            'Client': client,
            'Carrier': carrier,
            'Validity Date': validity_date,
            # Page Stopper and Table Name are only meaningful on header rows;
            # on data rows these fields are left blank
            'Page Stopper': item.get('PageStopper', '') if is_header else '',
            'Table Name': item.get('TableName', '') if is_header else '',
            'Weight From': weight_from,
            'Weight To': weight_to,
        }

        # For each zone column, read the value from the JSON item using the internal key
        # (e.g. "Zone1") but write it under the human-readable display label (e.g. "Zone 1").
        # On a header row this writes the zone name; on a data row this writes the price.
        for zone_key, zone_label in zone_column_names:
            row[zone_label] = item.get(zone_key, '')

        rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Excel sheet writers
# ---------------------------------------------------------------------------

def write_matrix_sheet(workbook, sheet_name, matrix_rows, category_specs, metadata):
    """
    Write the MainCosts tab to Excel with a special three-row header structure.

    WHY THREE HEADER ROWS?
    The MainCosts tab groups prices by cost category (e.g. "Documents", "Parcels").
    Each category has multiple weight columns (0.5 KG, 1 KG, 2 KG …).
    To make this readable, the header spans three rows:

      Row 1: Lane # | Origin | Destination | Service | Matrix zone | <-- Documents --> | <-- Parcels -->
      Row 2:        |        |             |         |             | Weight measure-KG |                |
      Row 3:        |        |             |         |             | 0.5 | 1 | 2 | 5   | 0.5 | 1 | 2 |
      Row 4+: actual data

    The category name in Row 1 is merged across all its weight columns.
    Data rows start at row 4.
    """
    if not matrix_rows:
        print(f"[WARN] No matrix data for {sheet_name}, skipping")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"[*] Creating {sheet_name} (Matrix) tab with {len(matrix_rows)} lanes...")
    ws = workbook.create_sheet(sheet_name)

    # Define the blue header style used for all three header rows
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # These five columns always appear first (left side of the sheet)
    fixed_cols = ['Lane #', 'Origin', 'Destination', 'Service', 'Matrix zone']
    num_fixed = len(fixed_cols)
    col = 1   # tracks the current column position as we build the header

    # --- Write the five fixed column names in Row 1 ---
    for c, name in enumerate(fixed_cols, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    col = num_fixed + 1   # move the column pointer past the fixed columns

    # --- Build the cost category column groups (Rows 1, 2, and 3) ---
    # For each cost category (e.g. "Documents"), we create a block of columns:
    #   - 1 extra column at the start of the block (for the "Weight measure - KG" label)
    #   - Then one column per weight breakpoint (0.5, 1, 2 …)
    category_start_cols = []   # stores position info for each category group

    for cost_cat_name, weight_unit, weights in category_specs:
        start_col = col   # remember where this category group starts

        # Row 2, first column of this group: write "Weight measure - KG"
        weight_measure_label = f"Weight measure - {weight_unit}" if weight_unit else "Weight measure"
        ws.cell(row=2, column=col, value=weight_measure_label)
        ws.cell(row=2, column=col).fill = header_fill
        ws.cell(row=2, column=col).font = header_font
        ws.cell(row=2, column=col).alignment = header_alignment
        col += 1

        # Row 2, remaining columns in this group: empty but styled (just the blue background)
        for _ in weights:
            ws.cell(row=2, column=col, value='')
            ws.cell(row=2, column=col).fill = header_fill
            col += 1

        # Row 3, first column of this group: empty (aligns with the "Weight measure" label above)
        ws.cell(row=3, column=start_col, value='')
        ws.cell(row=3, column=start_col).fill = header_fill
        col = start_col + 1

        # Row 3, remaining columns: write each weight breakpoint value (e.g. 0.5, 1, 2)
        for w in weights:
            ws.cell(row=3, column=col, value=w)
            ws.cell(row=3, column=col).fill = header_fill
            ws.cell(row=3, column=col).font = header_font
            ws.cell(row=3, column=col).alignment = header_alignment
            col += 1

        end_col = col - 1   # last column of this category group

        # Save the position info so we can write data rows correctly later
        category_start_cols.append((start_col, end_col, cost_cat_name, weight_unit, weights))

        # Row 1: merge all columns in this group into one cell and write the category name
        # e.g. "Documents" spans columns 6 to 10
        if start_col <= end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=cost_cat_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    total_cols = col - 1   # total number of columns in the sheet

    # Rows 2 and 3 under the five fixed columns: empty cells with the blue header fill
    # (so the header band looks continuous across the full width)
    for r in (2, 3):
        for c in range(1, num_fixed + 1):
            ws.cell(row=r, column=c, value='')
            ws.cell(row=r, column=c).fill = header_fill

    # --- Write the data rows starting at row 4 ---
    for row_idx, row_data in enumerate(matrix_rows, 4):
        col = 1

        # Write the five fixed columns (Lane #, Origin, Destination, Service, Matrix zone)
        for fc in fixed_cols:
            val = row_data.get(fc, '')
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fc == 'Lane #':
                cell.alignment = Alignment(horizontal="center")   # numbers look better centred
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            col += 1

        # Write the price columns for each cost category group.
        # The first column of each group is a spacer (empty).
        # The remaining columns each hold the price for one weight breakpoint.
        # The price is looked up from the row dict using the composite key (category, weight).
        for start_col, end_col, cost_cat_name, weight_unit, weights in category_start_cols:
            # First column of the group: spacer (empty)
            cell = ws.cell(row=row_idx, column=start_col, value='')
            cell.alignment = Alignment(horizontal="center")
            col = start_col + 1

            # One column per weight: look up the price and write it
            for w in weights:
                key = (cost_cat_name, w)   # e.g. ("Documents", "0.5")
                val = row_data.get(key, '')   # e.g. 10.50, or '' if no price for this weight
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = Alignment(horizontal="center")
                col += 1

    # --- Auto-size column widths ---
    # Sample the content of up to 53 rows (3 header rows + first 50 data rows)
    # to estimate a good column width.  Cap at 50 characters to avoid very wide columns.
    last_data_row = len(matrix_rows) + 3
    for c in range(1, total_cols + 1):
        col_letter = get_column_letter(c)
        max_len = 10   # minimum width
        for r in range(1, min(last_data_row + 1, 54)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Freeze the first three rows so the header stays visible when scrolling down
    ws.freeze_panes = "A4"
    # Add a filter dropdown to row 3 (the weight-breakpoint row) so users can filter
    ws.auto_filter.ref = f"A3:{get_column_letter(total_cols)}{last_data_row}"
    print(f"[OK] {sheet_name} (Matrix) tab created with {total_cols} columns")


def write_sheet(workbook, sheet_name, rows, metadata):
    """
    Write a standard flat-table Excel sheet (used for AddedRates, CountryZoning,
    AdditionalZoning, ZoningMatrix, AdditionalCostsPart1, AdditionalCostsPart2).

    This is the generic writer used for all tabs except MainCosts (which has its own
    special three-row header).  It produces a simple one-row header + data rows layout.

    COLUMN ORDERING:
    Columns are arranged in three groups, in this order:
      1. Priority columns  – always appear first in a fixed human-friendly sequence
                             (Client, Carrier, Validity Date, Country, Country Code, …)
      2. Weight columns    – columns whose name contains "KG", starts with "<=", or contains "-"
                             sorted numerically (0.5 KG before 1 KG before 2 KG)
      3. Zone columns      – "Zone 1", "Zone 2" … sorted numerically
         Other columns     – everything else, sorted alphabetically
    """
    if not rows:
        print(f"[WARN] No data for {sheet_name}, skipping")
        return
    
    print(f"[*] Creating {sheet_name} tab with {len(rows)} rows...")
    
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    ws = workbook.create_sheet(sheet_name)
    
    # Collect every column name that appears in any row (some rows may have extra fields)
    all_columns = set()
    for row in rows:
        all_columns.update(row.keys())
    
    # -----------------------------------------------------------------------
    # Step 1: Place the priority columns first.
    # These are the most important / most commonly used columns and should
    # always appear on the left side of the sheet.
    # -----------------------------------------------------------------------
    priority_cols = [
        'Client', 'Carrier', 'Validity Date',   # identity columns (always first)
        'Section', 'Service Type', 'Cost Category', 'Weight Unit', 'Zone',
        'Page Stopper', 'Table Name', 'Weight From', 'Weight To',
        'RateName', 'Country', 'Country Code', 'WeightFrom', 'WeightTo'
    ]
    
    columns = []
    for col in priority_cols:
        if col in all_columns:
            columns.append(col)
            all_columns.discard(col)   # remove from the remaining set so it doesn't appear twice
    
    # -----------------------------------------------------------------------
    # Step 2: From the remaining columns, separate weight columns from everything else.
    # Weight columns are identified by their name pattern:
    #   - Contains "KG"    e.g. "0.5 KG", "1 KG"
    #   - Starts with "<=" e.g. "<=0.5"
    #   - Contains "-"     e.g. "0-0.5"
    # -----------------------------------------------------------------------
    weight_cols = []
    other_cols = []
    
    for col in all_columns:
        if 'KG' in col or col.startswith('<=') or '-' in col:
            weight_cols.append(col)
        else:
            other_cols.append(col)
    
    # Sort weight columns numerically by the leading number
    # e.g. "0.5 KG", "1 KG", "2 KG" (not "0.5 KG", "2 KG", "1 KG")
    try:
        weight_cols_sorted = sorted(weight_cols, key=lambda x: float(x.split()[0]))
    except Exception:
        weight_cols_sorted = sorted(weight_cols)   # fallback: alphabetical

    # Sort "Zone N" columns numerically (Zone 1, Zone 2, Zone 10 …)
    # and sort all other columns alphabetically after them.
    def _other_col_sort_key(c):
        m = re.match(r'^Zone\s+(\d+)$', c, re.IGNORECASE)
        if m:
            return (0, int(m.group(1)))   # group 0: sort by zone number
        return (1, c)                      # group 1: sort alphabetically

    columns.extend(weight_cols_sorted)
    columns.extend(sorted(other_cols, key=_other_col_sort_key))
    
    # Define the blue header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write the header row (row 1) with the column names
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write the data rows starting at row 2.
    # For each row, look up the value for each column and write it to the correct cell.
    # If a row doesn't have a value for a column, write an empty string.
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, column in enumerate(columns, 1):
            value = row_data.get(column, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Short numeric/code values are centred; longer text values wrap inside the cell
            if column in ['Weight', 'Weight Unit', 'Section', 'Zone', 'Currency', 'Rate'] or 'KG' in column:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-size column widths by looking at the content of the first 50 data rows.
    # The width is capped between 10 and 50 characters to avoid extremes.
    for col_idx, column in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(column))   # start with the header name length as the minimum
        for row_idx in range(2, min(len(rows) + 2, 52)):   # sample up to 50 data rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Freeze the header row so column names stay visible when scrolling down
    ws.freeze_panes = "A2"
    # Add filter dropdowns to every column so users can filter/sort the data
    ws.auto_filter.ref = ws.dimensions
    
    print(f"[OK] {sheet_name} tab created with {len(columns)} columns")


# ---------------------------------------------------------------------------
# Accessorial Costs tab
# ---------------------------------------------------------------------------

# This list defines the exact columns and their order for the Accessorial Costs sheet.
# It is defined here as a constant so both the row-builder and the sheet-writer use
# the same column order without having to pass it around.
ACCESSORIAL_COSTS_COLUMNS = [
    'Original Cost Name',          # the cost name as it appears in the rate card PDF
    'Cost Type',                   # standardised type name (filled by fuzzy matching)
    'Cost Price',                  # the numeric price value
    'Currency',                    # e.g. EUR, USD
    'Rate by',                     # how the price is applied (e.g. per shipment, per kg)
    'Apply Over',                  # what the cost applies to (e.g. base freight)
    'Apply if',                    # condition under which the cost applies (left blank)
    'Additional info(Cost Code)',  # internal cost code from the rate card
    'Valid From',                  # start date of validity (taken from the rate card metadata)
    'Valid To',                    # end date of validity (not available; left blank)
    'Carrier',                     # carrier name
]


def _load_accessorial_cost_type_names(ref_path):
    """
    Read the list of approved/canonical cost type names from a reference file.

    PURPOSE:
    The rate card PDF uses its own names for costs (e.g. "Premium 9:00 Delivery").
    The business wants these mapped to standardised names from an approved list
    (e.g. "9:00 Service Fee").  This function loads that approved list.

    The reference file must have a column called 'Name'.  Supported file formats:
      - Excel (.xlsx or .xls)
      - CSV (.csv)

    Returns a deduplicated list of name strings, in the order they appear in the file.
    Returns an empty list [] if the file doesn't exist or has no 'Name' column.
    """
    ref_path = Path(ref_path)
    if not ref_path.exists():
        return []   # file not found; return empty list
    names = []
    try:
        if ref_path.suffix.lower() in ('.xlsx', '.xls'):
            import openpyxl
            # Open the Excel file in read-only mode (faster, doesn't load formulas)
            wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
            ws = wb.active   # use the first (active) sheet
            header = None
            name_col = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    # The first row is the header row.
                    # Find which column index has the heading "Name".
                    header = [str(c).strip() if c is not None else '' for c in row]
                    for i, h in enumerate(header):
                        if h == 'Name':
                            name_col = i   # remember the column index (0-based)
                            break
                    if name_col is None:
                        break   # no 'Name' column found; stop reading
                    continue    # move on to the data rows
                # For each data row, read the value in the 'Name' column
                if name_col is not None and name_col < len(row):
                    val = row[name_col]
                    if val is not None and str(val).strip():
                        names.append(str(val).strip())
            wb.close()
        elif ref_path.suffix.lower() == '.csv':
            import csv
            with open(ref_path, 'r', encoding='utf-8-sig', newline='') as f:
                reader = csv.reader(f)
                header = next(reader, None)   # read the first row as the header
                if header:
                    try:
                        name_col = header.index('Name')   # find the 'Name' column index
                    except ValueError:
                        name_col = None   # 'Name' column not found
                    if name_col is not None:
                        for row in reader:
                            if name_col < len(row) and row[name_col].strip():
                                names.append(row[name_col].strip())
        else:
            return []   # unsupported file format
    except Exception:
        return []   # something went wrong reading the file; return empty list

    # Remove duplicates while keeping the original order
    return list(dict.fromkeys(names))


def _token_set(text):
    """
    Break a text string into a set of individual words (tokens) in lowercase.
    Also handles time-like tokens such as "9:00" as a single token (not split at the colon).

    Example: "Premium 9:00 Delivery Fee" -> {"premium", "9:00", "delivery", "fee"}

    This is used by the fuzzy matching function to compare cost names word-by-word.
    """
    import re
    s = (text or '').lower().strip()
    # Match either "word:word" (time codes like 9:00) or plain alphanumeric words
    tokens = set(re.findall(r'[a-z0-9]+(?::[a-z0-9]+)?|[a-z]+', s))
    return tokens


def _best_match_cost_type(original_name, name_list, cutoff=0.4):
    """
    Find the best matching canonical cost type name for a given original cost name.

    WHY FUZZY MATCHING?
    The cost names in the rate card PDF (e.g. "Premium 9:00:") don't always match
    exactly the standardised names in the reference file (e.g. "9:00 Service Fee").
    We use a scoring system to find the closest match.

    HOW THE SCORE WORKS:
    For each candidate name in the reference list, we compute a combined score:
      score = character_similarity + token_overlap_bonus

      character_similarity: a 0-to-1 score from Python's difflib library that measures
                            how similar two strings look character by character.
                            e.g. "Premium 9:00" vs "9:00 Service Fee" -> ~0.35

      token_overlap_bonus:  an extra bonus (up to 0.4) for shared meaningful words.
                            "Meaningful" means the word is at least 2 characters long
                            or contains ":" (to catch time codes like "9:00").
                            e.g. both contain "9:00" -> bonus = 0.4

    The candidate with the highest combined score wins.
    If the winning score is below 0.3, we return '' (no match good enough).
    """
    if not original_name or not name_list:
        return ''
    original = str(original_name).strip()
    if not original:
        return ''

    orig_tokens = _token_set(original)   # break the original name into tokens
    best_score = -1.0
    best_name = ''

    for name in name_list:
        name_str = str(name).strip()
        if not name_str:
            continue

        # Measure character-level similarity (0 = completely different, 1 = identical)
        char_ratio = difflib.SequenceMatcher(None, original.lower(), name_str.lower()).ratio()

        name_tokens = _token_set(name_str)
        shared = orig_tokens & name_tokens   # tokens that appear in both names

        # Calculate the token overlap bonus.
        # Only count "meaningful" tokens (length >= 2 or contains ':' for time codes).
        meaningful_orig = {t for t in orig_tokens if len(t) >= 2 or ':' in t}
        if meaningful_orig:
            # What fraction of the original's meaningful tokens also appear in the candidate?
            # Multiply by 0.4 so the bonus can add at most 0.4 to the score.
            token_bonus = (len(shared & meaningful_orig) / len(meaningful_orig)) * 0.4
        else:
            token_bonus = 0.0

        score = char_ratio + token_bonus

        # Keep track of the highest-scoring candidate seen so far
        if score > best_score:
            best_score = score
            best_name = name_str

    # Only return a match if the score is high enough to be trustworthy
    return best_name if best_score >= 0.3 else ''


def build_accessorial_costs_rows(additional_costs_1, additional_costs_2, metadata, cost_type_ref_path=None, accessorial_folder=None):
    """
    Build the rows for the "Accessorial Costs" Excel tab.

    WHAT ARE ACCESSORIAL COSTS?
    These are extra charges on top of the base shipping rate, such as:
    fuel surcharges, remote area fees, signature fees, Saturday delivery fees, etc.
    They come from two sections of the JSON: AdditionalCostsPart1 and AdditionalCostsPart2.

    WHAT THIS FUNCTION DOES:
    1. Converts every item from both parts into a row matching ACCESSORIAL_COSTS_COLUMNS.
    2. Tries to fill the "Cost Type" column by matching each "Original Cost Name" against
       a reference file of approved cost type names for this client.
       - The reference file is found by looking in accessorial_folder for a file whose
         filename contains the client name (e.g. "Acme_Accessorial_Costs.xlsx" for client "Acme").
       - Matching is done by _best_match_cost_type() (fuzzy/approximate matching).
       - If no reference file is found, Cost Type is left blank.

    Returns: (list_of_rows, path_of_reference_file_used_or_None)
    """
    carrier = (metadata.get('carrier') or '').replace('\n', ' ')
    validity_date = (metadata.get('validity_date') or '')

    def item_to_row(item):
        """Convert one JSON cost item into a row dict matching ACCESSORIAL_COSTS_COLUMNS."""
        # The price field can be named either "CostPrice" or "CostAmount" depending on the source
        cost_price = item.get('CostPrice') or item.get('CostAmount') or ''
        return {
            'Original Cost Name': item.get('CostName', ''),      # name as in the PDF
            'Cost Type': '',                                       # filled later by fuzzy matching
            'Cost Price': cost_price,
            'Currency': item.get('CostCurrency', ''),
            'Rate by': item.get('PriceMechanism', ''),            # e.g. "Per Shipment", "Per KG"
            'Apply Over': item.get('ApplyTo', ''),                # e.g. "Base Freight"
            'Apply if': '',                                        # not in source JSON; left blank
            'Additional info(Cost Code)': item.get('CostCode', ''),
            'Valid From': validity_date,
            'Valid To': '',                                        # not in source JSON; left blank
            'Carrier': carrier,
        }

    # Combine AdditionalCostsPart1 and AdditionalCostsPart2 into one flat list.
    # Part1 rows come first, then Part2 rows.
    rows = []
    for item in additional_costs_1 or []:
        rows.append(item_to_row(item))
    for item in additional_costs_2 or []:
        rows.append(item_to_row(item))

    # -----------------------------------------------------------------------
    # Find the reference file for Cost Type fuzzy matching.
    #
    # If no explicit file path was provided, search the accessorial_folder for a
    # file whose name contains the client name (case-insensitive).
    # e.g. client "Acme" matches file "Acme_Accessorial_Costs.xlsx"
    # -----------------------------------------------------------------------
    if cost_type_ref_path is None and accessorial_folder:
        accessorial_dir = Path(accessorial_folder)
        client = (metadata.get('client') or '').strip()
        ext_order = ('.xlsx', '.xls', '.csv')   # preference order: xlsx first, then xls, then csv
        print("[*] Accessorial Cost Type mapping: client from JSON -> look in folder for file whose name contains client (case-insensitive) -> use that file's 'Name' column to fill Cost Type for each row.")
        print(f"    Folder: {accessorial_dir}")
        print(f"    Client: {client or '(none)'}")

        if client and accessorial_dir.exists() and accessorial_dir.is_dir():
            client_lower = client.lower()
            # List all files in the folder whose name contains the client name
            candidates = [
                p for p in accessorial_dir.iterdir()
                if p.is_file()
                and p.suffix.lower() in ext_order
                and client_lower in p.stem.lower()   # p.stem is filename without extension
            ]
            if candidates:
                # If multiple files match, prefer xlsx over xls over csv
                cost_type_ref_path = min(
                    candidates,
                    key=lambda p: ext_order.index(p.suffix.lower()) if p.suffix.lower() in ext_order else 99,
                )
                print(f"[*] Accessorial cost mapping: using file (client '{client}' in filename) {cost_type_ref_path.name}")
                print(f"    Using file: {cost_type_ref_path.name}  (path: {cost_type_ref_path})")
            else:
                print(f"[*] Accessorial cost mapping: no file with client '{client}' in name in {accessorial_dir}")
        elif not client:
            print(f"[*] Accessorial cost mapping: no client in metadata, Cost Type left empty")
        elif not accessorial_dir.exists() or not accessorial_dir.is_dir():
            print(f"[*] Accessorial cost mapping: folder not found {accessorial_dir}, Cost Type left empty")

    if cost_type_ref_path:
        # Load the list of approved cost type names from the reference file
        name_list = _load_accessorial_cost_type_names(cost_type_ref_path)
        if name_list:
            # For every row, try to find the best matching canonical name
            # and write it into the 'Cost Type' column
            for row in rows:
                original = row.get('Original Cost Name', '')
                row['Cost Type'] = _best_match_cost_type(original, name_list)
            print(f"[*] Accessorial Cost Type: filled from {cost_type_ref_path.name} ({len(name_list)} cost types, {len(rows)} rows)")
        else:
            print(f"[*] Accessorial Cost Type: file {cost_type_ref_path.name} has no 'Name' column or is empty, Cost Type left blank")

    return rows, cost_type_ref_path


def write_accessorial_sheet(workbook, sheet_name, rows):
    """
    Write the Accessorial Costs tab to Excel.

    This is a simplified version of write_sheet() that uses the fixed column order
    defined in ACCESSORIAL_COSTS_COLUMNS instead of dynamically determining columns.
    The column order is fixed because the Accessorial Costs tab has a specific agreed layout.
    """
    if not rows:
        print(f"[WARN] No data for {sheet_name}, skipping")
        return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    print(f"[*] Creating {sheet_name} tab with {len(rows)} rows...")
    ws = workbook.create_sheet(sheet_name)
    columns = ACCESSORIAL_COSTS_COLUMNS   # use the fixed column list defined at the top of this section

    # Define the blue header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write the header row (row 1) with the fixed column names
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write the data rows starting at row 2.
    # All cells use wrap_text so long cost names are readable without widening the column too much.
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, column in enumerate(columns, 1):
            value = row_data.get(column, '')   # empty string if this row has no value for this column
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size columns by sampling up to 100 data rows (more than write_sheet's 50,
    # because cost names can be long and we want to capture outliers)
    for col_idx, column in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(column))   # start with the header name length
        for row_idx in range(2, min(len(rows) + 2, 102)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)

    # Freeze the header row and add filter dropdowns
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    print(f"[OK] {sheet_name} tab created with {len(columns)} columns")


# ---------------------------------------------------------------------------
# Workbook orchestrator
# ---------------------------------------------------------------------------

def save_to_excel(data, output_path, accessorial_folder=None):
    """
    The main function that builds the complete Excel workbook and saves it to disk.

    This function is the "conductor" – it calls all the other functions in the right
    order and assembles their output into a single multi-tab Excel file.

    TABS CREATED (in order):
      1  Metadata              – basic info: client, carrier, validity date, filename
      2  MainCosts             – the main pricing table (one row per lane, all cost
                                 categories merged, letter zones expanded to real O/D pairs)
      3  AddedRates            – supplemental rate tables (e.g. fuel surcharge by weight/zone)
      4  AdditionalCostsPart1  – first batch of extra charges from the rate card
      5  CountryZoning         – which countries belong to which zone (with ISO codes added)
      6  AdditionalZoning      – additional zoning rules (if present)
      7  ZoningMatrix          – the raw origin/destination zone matrix (for reference)
      8  AdditionalCostsPart2  – second batch of extra charges from the rate card
      9  Accessorial Costs     – combined view of Part1 + Part2 with standardised Cost Types

    PARAMETERS:
      data               – the full JSON dictionary loaded from extracted_data.json
      output_path        – where to save the .xlsx file
      accessorial_folder – folder containing client-specific reference files for Cost Type matching

    RETURNS: the path of the accessorial reference file used (or None if none was found)
    """
    print(f"[*] Creating Excel file: {output_path}")
    
    try:
        import openpyxl   # the library used to create and write Excel files
    except ImportError:
        print("[ERROR] openpyxl not installed!")
        print("        To install: pip install openpyxl")
        raise
    
    try:
        # Create a new empty workbook.
        # openpyxl always creates one default sheet; we remove it immediately
        # because we will add our own sheets with specific names.
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Extract the metadata section from the JSON (carrier name, validity date, etc.)
        metadata = data.get('metadata', {})
        
        # -----------------------------------------------------------------------
        # Tab 1: Metadata
        # Simple two-column table showing what rate card this file came from.
        # -----------------------------------------------------------------------
        create_metadata_sheet(wb, metadata)
        
        # -----------------------------------------------------------------------
        # Tab 2: MainCosts
        # The most complex tab.  Two-step process:
        #   Step A: build_matrix_main_costs() merges all cost categories into one row per lane
        #   Step B: expand_main_costs_lanes_by_zoning() replaces letter zones (A, B …)
        #           with real Origin/Destination pairs from the ZoningMatrix
        # -----------------------------------------------------------------------
        main_costs_data = data.get('MainCosts', [])
        zoning_matrix = data.get('ZoningMatrix', [])
        if main_costs_data:
            matrix_rows, category_specs = build_matrix_main_costs(main_costs_data, metadata)
            if zoning_matrix:
                # Only expand if there is ZoningMatrix data to expand with
                matrix_rows = expand_main_costs_lanes_by_zoning(matrix_rows, zoning_matrix)
            write_matrix_sheet(wb, "MainCosts", matrix_rows, category_specs, metadata)
        
        # -----------------------------------------------------------------------
        # Tab 3: AddedRates
        # Supplemental rate tables (e.g. fuel surcharge, remote area surcharge).
        # The JSON has interleaved header and data rows; pivot_added_rates() untangles them.
        # -----------------------------------------------------------------------
        added_rates = data.get('AddedRates', [])
        if added_rates:
            added_rates_rows = pivot_added_rates(added_rates, metadata)
            write_sheet(wb, "AddedRates", added_rates_rows, metadata)
        
        # -----------------------------------------------------------------------
        # Tab 4: AdditionalCostsPart1
        # First batch of extra charges.  No special transformation needed;
        # flatten_array_data() just prepends the three identity columns.
        # -----------------------------------------------------------------------
        additional_costs_1 = data.get('AdditionalCostsPart1', [])
        if additional_costs_1:
            additional_costs_1_rows = flatten_array_data(additional_costs_1, metadata, 'AdditionalCostsPart1')
            write_sheet(wb, "AdditionalCostsPart1", additional_costs_1_rows, metadata)
        
        # -----------------------------------------------------------------------
        # Tab 5: CountryZoning
        # Maps countries to delivery zones.
        # flatten_array_data() applies two extra enrichment steps for this tab:
        #   - Forward-fill empty RateName cells
        #   - Add a Country Code column (ISO 2-letter codes)
        # -----------------------------------------------------------------------
        country_zoning = data.get('CountryZoning', [])
        if country_zoning:
            country_zoning_rows = flatten_array_data(country_zoning, metadata, 'CountryZoning')
            write_sheet(wb, "CountryZoning", country_zoning_rows, metadata)
        
        # -----------------------------------------------------------------------
        # Tab 6: AdditionalZoning
        # Extra zoning rules (if present in the rate card).  Flat pass-through.
        # -----------------------------------------------------------------------
        additional_zoning = data.get('AdditionalZoning', [])
        if additional_zoning:
            additional_zoning_rows = flatten_array_data(additional_zoning, metadata, 'AdditionalZoning')
            write_sheet(wb, "AdditionalZoning", additional_zoning_rows, metadata)
        
        # -----------------------------------------------------------------------
        # Tab 7: ZoningMatrix
        # The raw origin/destination zone matrix (the same data used in Tab 2 to
        # expand letter zones).  Shown here for reference/audit purposes.
        # -----------------------------------------------------------------------
        zoning_matrix = data.get('ZoningMatrix', [])
        if zoning_matrix:
            zoning_matrix_rows = flatten_array_data(zoning_matrix, metadata, 'ZoningMatrix')
            write_sheet(wb, "ZoningMatrix", zoning_matrix_rows, metadata)
        
        # -----------------------------------------------------------------------
        # Tab 8: AdditionalCostsPart2
        # Second batch of extra charges.  Flat pass-through.
        # -----------------------------------------------------------------------
        additional_costs_2 = data.get('AdditionalCostsPart2', [])
        if additional_costs_2:
            additional_costs_2_rows = flatten_array_data(additional_costs_2, metadata, 'AdditionalCostsPart2')
            write_sheet(wb, "AdditionalCostsPart2", additional_costs_2_rows, metadata)
        
        # -----------------------------------------------------------------------
        # Tab 9: Accessorial Costs
        # Combines Part1 and Part2 into one clean table with standardised Cost Types.
        # build_accessorial_costs_rows() handles the fuzzy matching of cost type names.
        # -----------------------------------------------------------------------
        accessorial_rows, accessorial_file_used = build_accessorial_costs_rows(
            data.get('AdditionalCostsPart1', []),
            data.get('AdditionalCostsPart2', []),
            metadata,
            accessorial_folder=accessorial_folder,
        )
        if accessorial_rows:
            write_accessorial_sheet(wb, "Accessorial Costs", accessorial_rows)
        
        # Save the finished workbook to the output file path
        wb.save(output_path)
        
        # Report the file size so the user can see the output was created successfully
        file_size = os.path.getsize(output_path)
        file_size_kb = file_size / 1024
        
        print(f"[OK] Excel file saved successfully")
        print(f"  - Tabs: {len(wb.sheetnames)}")
        print(f"  - File size: {file_size_kb:.2f} KB")
        
        # Return the path of the accessorial reference file that was used (or None)
        return str(accessorial_file_used) if accessorial_file_used else None
        
    except Exception as e:
        print(f"[ERROR] Failed to save Excel: {e}")
        raise


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    """
    The starting point when this script is run directly from the command line.

    This function:
      1. Defines the input and output file paths.
      2. Makes sure the output folder exists (creates it if needed).
      3. Loads the extracted JSON data from disk.
      4. Calls save_to_excel() to build and save the Excel workbook.
      5. Prints a summary of what was created.

    If anything goes wrong at any step, an error message is printed and the
    program stops with a non-zero exit code.
    """
    print("=" * 60)
    print("DHL RATE CARD EXCEL GENERATOR")
    print("=" * 60)
    print()
    
    # Define where to read from and where to write to
    input_file = 'processing/extracted_data.json'   # the JSON produced by the extraction pipeline
    output_dir = 'output'                            # folder where the Excel file will be saved
    output_file = os.path.join(output_dir, 'DHL_Rate_Cards.xlsx')   # full output path
    
    try:
        # Create the output folder if it doesn't already exist.
        # parents=True means it will also create any missing parent folders.
        # exist_ok=True means it won't fail if the folder already exists.
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        print(f"[OK] Output directory ready: {output_dir}")
        print()
        
        # Step 1: Read the JSON file from disk into a Python dictionary
        print("Step 1: Loading extracted data...")
        data = load_extracted_data(input_file)
        print()
        
        # Step 2: Build the Excel workbook and save it
        print("Step 2: Creating multi-tab Excel file...")
        save_to_excel(data, output_file)
        print()
        
        # Print a success summary showing which tabs were created and how many rows each has
        print("=" * 60)
        print("[SUCCESS] EXCEL GENERATION COMPLETE")
        print("=" * 60)
        print(f"Output file: {output_file}")
        print()
        print("Tabs created:")
        print("  1. Metadata (Carrier, Validity info)")
        
        # The JSON may contain a 'statistics' section with row counts; use it for the summary
        stats = data.get('statistics', {})
        if stats.get('MainCosts_sections', 0) > 0:
            print(f"  2. MainCosts ({stats.get('MainCosts_rows', 0)} pricing rows)")
        if stats.get('AddedRates_rows', 0) > 0:
            print(f"  3. AddedRates ({stats.get('AddedRates_rows', 0)} rows)")
        if stats.get('AdditionalCostsPart1_rows', 0) > 0:
            print(f"  4. AdditionalCostsPart1 ({stats.get('AdditionalCostsPart1_rows', 0)} rows)")
        if stats.get('CountryZoning_rows', 0) > 0:
            print(f"  5. CountryZoning ({stats.get('CountryZoning_rows', 0)} rows)")
        if stats.get('AdditionalZoning_rows', 0) > 0:
            print(f"  6. AdditionalZoning ({stats.get('AdditionalZoning_rows', 0)} rows)")
        if stats.get('ZoningMatrix_rows', 0) > 0:
            print(f"  7. ZoningMatrix ({stats.get('ZoningMatrix_rows', 0)} rows)")
        if stats.get('AdditionalCostsPart2_rows', 0) > 0:
            print(f"  8. AdditionalCostsPart2 ({stats.get('AdditionalCostsPart2_rows', 0)} rows)")
        # For Accessorial Costs, count directly from the data (not from statistics)
        acc_count = len(data.get('AdditionalCostsPart1', [])) + len(data.get('AdditionalCostsPart2', []))
        if acc_count > 0:
            print(f"  9. Accessorial Costs ({acc_count} rows)")
        print()
        
    except Exception as e:
        # If anything went wrong, print a clear failure message before re-raising the error
        print()
        print("=" * 60)
        print("[FAILED] EXCEL GENERATION FAILED")
        print("=" * 60)
        print(f"Error: {e}")
        print()
        raise


# This block only runs when the script is executed directly (e.g. "python create_table.py").
# It does NOT run when this file is imported as a module by another script.
if __name__ == "__main__":
    main()
